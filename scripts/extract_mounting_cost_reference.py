from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_PATH = REPO_ROOT / 'src/domain/mounting-cost/model/mounting-cost-reference.generated.ts'
WORKBOOK_ENV_VAR = 'MOUNTING_REFERENCE_WORKBOOK'
SECTION_NUMBERS = tuple(range(1, 14))
TIER_ORDER = ['tier-1', 'tier-2', 'tier-3', 'tier-4']
REGION_KEYS = [
    'chelyabinsk-oblast',
    'sverdlovsk-tyumen',
    'hmao',
    'yanao',
    'volga-fo',
    'south-fo',
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Extract mounting cost reference data from matrix workbook.'
    )
    parser.add_argument('--input', help='Path to source XLSX file')
    parser.add_argument('--output', help='Path to output generated TS file')
    return parser.parse_args()


def discover_workbook_candidates() -> list[Path]:
    candidates: list[Path] = []
    env_path = os.environ.get(WORKBOOK_ENV_VAR)
    if env_path:
        candidates.append(Path(env_path))

    for base in (REPO_ROOT, REPO_ROOT.parent, Path('C:/montaz')):
        if not base.exists():
            continue
        for path in sorted(base.glob('*.xlsx')):
            if path.name.startswith('~$'):
                continue
            candidates.append(path)

    unique: list[Path] = []
    seen: set[Path] = set()
    for path in candidates:
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        unique.append(path)
    return unique


def first_existing_path(paths: list[Path]) -> Path | None:
    for path in paths:
        if path.exists() and path.is_file():
            return path
    return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def normalize_text_lower(value: Any) -> str:
    return normalize_text(value).lower()


def coerce_float(value: Any) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip().replace(',', '.')
        if normalized:
            return float(normalized)
    raise ValueError(f'Expected numeric value, got {value!r}')


def coerce_int(value: Any) -> int | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(float(value))
    if isinstance(value, str):
        normalized = value.strip().replace(',', '.')
        if re.fullmatch(r'-?\d+(?:\.\d+)?', normalized):
            return int(float(normalized))
    return None


def parse_section_number_from_title(title: str) -> int | None:
    match = re.match(r'^\s*(\d{1,2})\s*[\.)]?', title)
    if not match:
        return None
    parsed = int(match.group(1))
    if parsed in SECTION_NUMBERS:
        return parsed
    return None


def detect_section_header_number(
    section_cell: Any,
    subindex_cell: Any,
    title_cell: Any,
) -> int | None:
    title = normalize_text(title_cell)
    if not title or coerce_int(subindex_cell) is not None:
        return None
    if re.fullmatch(r'\d+(?:[.,]\d+)?', title):
        return None

    by_title = parse_section_number_from_title(title)
    if by_title is not None:
        return by_title

    by_cell = coerce_int(section_cell)
    if by_cell in SECTION_NUMBERS:
        return by_cell
    return None


def parse_tier_upper_bound(note_text: str) -> float | None:
    if not note_text:
        return None

    normalized = note_text.lower().replace(' ', '')
    if 'свыше' in normalized or normalized.startswith('>'):
        return None

    match = re.search(r'(\d+(?:[.,]\d+)?)', normalized)
    if not match:
        return None

    return float(match.group(1).replace(',', '.'))


def extract_sections_and_priced_rows(ws) -> tuple[dict[int, str], list[dict[str, Any]]]:
    section_titles: dict[int, str] = {}
    rows: list[dict[str, Any]] = []
    current_section_number: int | None = None

    for row in range(1, ws.max_row + 1):
        section_cell = ws.cell(row=row, column=1).value
        subindex_cell = ws.cell(row=row, column=2).value
        title_cell = ws.cell(row=row, column=3).value

        detected_section = detect_section_header_number(section_cell, subindex_cell, title_cell)
        if detected_section is not None:
            current_section_number = detected_section
            section_titles.setdefault(detected_section, normalize_text(title_cell))

        unit_price = ws.cell(row=row, column=9).value
        subindex = coerce_int(subindex_cell)
        if subindex is None or unit_price is None:
            continue

        section_number = (
            current_section_number
            if current_section_number in SECTION_NUMBERS
            else coerce_int(section_cell)
        )
        if section_number not in SECTION_NUMBERS:
            continue

        note = normalize_text(ws.cell(row=row, column=11).value)
        rows.append(
            {
                'sectionNumber': section_number,
                'subindex': subindex,
                'itemCode': f'{section_number}.{subindex}',
                'itemTitle': normalize_text(title_cell),
                'unit': normalize_text(ws.cell(row=row, column=4).value),
                'basis': normalize_text(ws.cell(row=row, column=5).value),
                'unitPriceRubPerUnit': round(coerce_float(unit_price)),
                'note': note,
                'tierUpperBoundQuantity': parse_tier_upper_bound(note),
            }
        )

    return section_titles, rows


def sort_rows_for_tiers(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sorted_rows = sorted(
        rows,
        key=lambda item: (
            item['tierUpperBoundQuantity'] is None,
            item['tierUpperBoundQuantity']
            if item['tierUpperBoundQuantity'] is not None
            else float('inf'),
            item['subindex'],
        ),
    )
    if len(sorted_rows) < 4:
        raise ValueError(f'Expected at least 4 tier rows, got {len(sorted_rows)}')
    return sorted_rows[:4]


def build_tiered_reference(
    section_key: str,
    section_title: str,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    tier_rows = sort_rows_for_tiers(rows)
    base_row = tier_rows[0]

    return {
        'sectionKey': section_key,
        'sectionTitle': section_title,
        'itemCode': base_row['itemCode'],
        'itemTitle': base_row['itemTitle'],
        'unit': base_row['unit'],
        'basis': base_row['basis'],
        'tierUpperBoundQuantity': {
            tier: tier_rows[index]['tierUpperBoundQuantity']
            for index, tier in enumerate(TIER_ORDER)
        },
        'unitPriceByTierRubPerUnit': {
            tier: tier_rows[index]['unitPriceRubPerUnit']
            for index, tier in enumerate(TIER_ORDER)
        },
    }


def build_fixed_reference(
    section_key: str,
    section_title: str,
    row: dict[str, Any],
) -> dict[str, Any]:
    return {
        'sectionKey': section_key,
        'sectionTitle': section_title,
        'itemCode': row['itemCode'],
        'itemTitle': row['itemTitle'],
        'unit': row['unit'],
        'basis': row['basis'],
        'unitPriceRubPerUnit': row['unitPriceRubPerUnit'],
    }


def require_rows(
    rows: list[dict[str, Any]],
    predicate: Callable[[dict[str, Any]], bool],
    *,
    context: str,
) -> list[dict[str, Any]]:
    filtered = [row for row in rows if predicate(row)]
    if not filtered:
        raise ValueError(f'No priced rows matched for {context}')
    return filtered


def require_row_by_subindex(
    rows: list[dict[str, Any]],
    subindex: int,
    *,
    context: str,
) -> dict[str, Any]:
    for row in rows:
        if row['subindex'] == subindex:
            return row
    raise ValueError(f'No priced row with subindex={subindex} for {context}')


def extract_region_coefficients(ws) -> dict[str, float]:
    runs: list[list[tuple[int, float]]] = []
    current_run: list[tuple[int, float]] = []

    for row in range(1, ws.max_row + 1):
        raw_name = normalize_text(ws.cell(row=row, column=3).value)
        raw_coeff = ws.cell(row=row, column=4).value

        coeff: float | None = None
        try:
            if raw_coeff is not None:
                parsed = coerce_float(raw_coeff)
                if 0.5 <= parsed <= 2.5 and raw_name:
                    coeff = parsed
        except ValueError:
            coeff = None

        if coeff is None:
            if current_run:
                runs.append(current_run)
                current_run = []
            continue

        current_run.append((row, coeff))

    if current_run:
        runs.append(current_run)

    candidate_runs = [run for run in runs if len(run) >= len(REGION_KEYS)]
    if not candidate_runs:
        raise ValueError('Region coefficients block was not found')

    selected_run = max(candidate_runs, key=lambda run: run[-1][0])
    selected = selected_run[: len(REGION_KEYS)]

    coefficients: dict[str, float] = {}
    for key, (_, coeff) in zip(REGION_KEYS, selected):
        coefficients[key] = coeff

    return coefficients


def build_reference_data(ws) -> dict[str, Any]:
    section_titles, priced_rows = extract_sections_and_priced_rows(ws)

    section_rows: dict[int, list[dict[str, Any]]] = {}
    for section_number in SECTION_NUMBERS:
        rows = [row for row in priced_rows if row['sectionNumber'] == section_number]
        if not rows:
            raise ValueError(f'No priced rows found for section {section_number}')
        section_rows[section_number] = rows

    missing_titles = [number for number in SECTION_NUMBERS if number not in section_titles]
    if missing_titles:
        raise ValueError(f'Section titles are missing in workbook: {missing_titles}')

    metal_lstk_rows = require_rows(
        section_rows[6],
        lambda row: row['subindex'] <= 4,
        context='section 6 LSTK',
    )
    metal_sort_rows = require_rows(
        section_rows[6],
        lambda row: row['subindex'] >= 5,
        context='section 6 black steel',
    )
    floor_150_rows = require_rows(
        section_rows[5],
        lambda row: '150' in normalize_text_lower(row['itemTitle']),
        context='section 5 floors 150',
    )
    floor_200_rows = require_rows(
        section_rows[5],
        lambda row: '200' in normalize_text_lower(row['itemTitle']),
        context='section 5 floors 200',
    )
    roof_panels_rows = require_rows(
        section_rows[8],
        lambda row: row['subindex'] <= 4,
        context='section 8 roof panels',
    )
    roof_fence_and_snow_rows = require_rows(
        section_rows[8],
        lambda row: 5 <= row['subindex'] <= 8,
        context='section 8 roof fence and snow',
    )
    roof_drainage_rows = require_rows(
        section_rows[8],
        lambda row: 9 <= row['subindex'] <= 12,
        context='section 8 roof drainage',
    )

    section_6_lstk = build_tiered_reference(
        'section-6-metal-lstk',
        section_titles[6],
        metal_lstk_rows,
    )
    section_6_sort = build_tiered_reference(
        'section-6-metal-sort',
        section_titles[6],
        metal_sort_rows,
    )
    section_3 = build_tiered_reference(
        'section-3-earthworks',
        section_titles[3],
        section_rows[3],
    )
    section_5_150 = build_tiered_reference(
        'section-5-floors-150',
        section_titles[5],
        floor_150_rows,
    )
    section_5_200 = build_tiered_reference(
        'section-5-floors-200',
        section_titles[5],
        floor_200_rows,
    )
    section_7 = build_tiered_reference(
        'section-7-walls',
        section_titles[7],
        section_rows[7],
    )
    section_8_roof = build_tiered_reference(
        'section-8-roof-panels',
        section_titles[8],
        roof_panels_rows,
    )

    region_coefficients = extract_region_coefficients(ws)

    return {
        'snapshot': {
            'sourceWorkbook': 'Матрица 2026 СМР.xlsx',
            'sourceSheets': [ws.title],
            'status': 'parity-verified',
            'note': 'Generated from matrix sections 1..13 and region coefficients.',
        },
        'section1Geology': build_tiered_reference(
            'section-1-geology',
            section_titles[1],
            section_rows[1],
        ),
        'section2Project': build_tiered_reference(
            'section-2-project',
            section_titles[2],
            section_rows[2],
        ),
        'section3Earthworks': section_3,
        'section4Concrete': build_tiered_reference(
            'section-4-concrete',
            section_titles[4],
            section_rows[4],
        ),
        'section5FloorsByType': {
            'slab-150': section_5_150,
            'slab-200': section_5_200,
        },
        'section6MetalByFrameType': {
            'lstk': section_6_lstk,
            'sort': section_6_sort,
        },
        'section7Walls': section_7,
        'section8RoofPanels': section_8_roof,
        'section8RoofFenceAndSnowGuards': build_tiered_reference(
            'section-8-roof-fence-and-snow',
            section_titles[8],
            roof_fence_and_snow_rows,
        ),
        'section8RoofDrainage': build_tiered_reference(
            'section-8-roof-drainage',
            section_titles[8],
            roof_drainage_rows,
        ),
        'section9Openings': {
            'doubleDoor': build_fixed_reference(
                'section-9-openings-double-door',
                section_titles[9],
                require_row_by_subindex(section_rows[9], 1, context='section 9 openings'),
            ),
            'singleDoor': build_fixed_reference(
                'section-9-openings-single-door',
                section_titles[9],
                require_row_by_subindex(section_rows[9], 2, context='section 9 openings'),
            ),
            'entranceBlock': build_fixed_reference(
                'section-9-openings-entrance-block',
                section_titles[9],
                require_row_by_subindex(section_rows[9], 3, context='section 9 openings'),
            ),
            'tambourDoor': build_fixed_reference(
                'section-9-openings-tambour-door',
                section_titles[9],
                require_row_by_subindex(section_rows[9], 4, context='section 9 openings'),
            ),
            'windows': build_fixed_reference(
                'section-9-openings-windows',
                section_titles[9],
                require_row_by_subindex(section_rows[9], 5, context='section 9 openings'),
            ),
            'gates': build_fixed_reference(
                'section-9-openings-gates',
                section_titles[9],
                require_row_by_subindex(section_rows[9], 6, context='section 9 openings'),
            ),
        },
        'section10WaterSewer': build_tiered_reference(
            'section-10-water-sewer',
            section_titles[10],
            section_rows[10],
        ),
        'section11Heating': build_tiered_reference(
            'section-11-heating',
            section_titles[11],
            section_rows[11],
        ),
        'section12Electrical': build_tiered_reference(
            'section-12-electrical',
            section_titles[12],
            section_rows[12],
        ),
        'section13AdditionalWorks': build_fixed_reference(
            'section-13-additional-works',
            section_titles[13],
            require_row_by_subindex(section_rows[13], 1, context='section 13 additional works'),
        ),
        # Backward-compatible aliases for existing code paths.
        'metalStructuresByFrameType': {
            'lstk': section_6_lstk,
            'sort': section_6_sort,
        },
        'earthworks': section_3,
        'floorsByType': {
            'slab-150': section_5_150,
            'slab-200': section_5_200,
        },
        'walls': section_7,
        'roof': section_8_roof,
        'regionCoefficientByKey': region_coefficients,
    }


def render_module(reference_data: dict[str, Any], workbook_name: str, sheet_name: str) -> str:
    serialized = json.dumps(reference_data, ensure_ascii=False, indent=2)

    return (
        f'// Auto-generated from workbook "{workbook_name}" (sheet "{sheet_name}").\n'
        f'// Run `npm run generate:mounting-ref` to refresh.\n\n'
        f'export const mountingCostReference = {serialized} as const;\n'
    )


def main() -> int:
    args = parse_args()

    workbook_path = (
        Path(args.input)
        if args.input
        else first_existing_path(discover_workbook_candidates())
    )
    if workbook_path is None:
        raise FileNotFoundError(
            'Workbook not found. Set --input or MOUNTING_REFERENCE_WORKBOOK, '
            'or place matrix file in project/work parent/C:/montaz.'
        )

    output_path = Path(args.output) if args.output else DEFAULT_OUTPUT_PATH

    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    reference_data = build_reference_data(sheet)
    rendered = render_module(reference_data, workbook_path.name, sheet.title)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(rendered, encoding='utf-8')

    print(f'Generated {output_path}')
    print(f'Workbook: {workbook_path}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
